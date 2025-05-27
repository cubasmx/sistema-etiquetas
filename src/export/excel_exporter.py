from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any
import os

class ExcelExporter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Lista de Materiales"
        self.current_row = 1
        
        # Estilos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Estilo para niveles
        self.level_styles = {}
        for i in range(5):  # Soportar hasta 5 niveles de profundidad
            style = NamedStyle(name=f'level_{i}')
            style.font = Font(bold=True if i == 0 else False)
            style.fill = PatternFill(
                start_color="E6E6E6" if i % 2 == 0 else "FFFFFF",
                end_color="E6E6E6" if i % 2 == 0 else "FFFFFF",
                fill_type="solid"
            )
            self.level_styles[i] = style
            self.wb.add_named_style(style)
        
        # Estilo para operaciones
        self.operation_style = NamedStyle(name='operation')
        self.operation_style.font = Font(italic=True)
        self.operation_style.fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )
        self.wb.add_named_style(self.operation_style)
    
    def write_bom_line(self, line: Dict[str, Any], parent_qty: float = 1.0):
        """
        Escribe una línea de la BOM y sus sub-BOMs si existen
        Args:
            line: Línea de la BOM
            parent_qty: Cantidad del producto padre para cálculos acumulados
        """
        level = line.get('level', 0)
        style = self.level_styles[min(level, 4)]  # Limitar a 5 niveles de estilo
        
        # Calcular la cantidad acumulada
        qty = line['product_qty'] * parent_qty
        
        # Añadir indentación según el nivel
        indent = "    " * level
        
        # Código
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = line['product_id'][1].split(']')[0].strip('[') if '][' in line['product_id'][1] else ''
        cell.border = self.border
        cell.alignment = Alignment(horizontal="center")
        cell.style = style
        
        # Componente
        cell = self.ws.cell(row=self.current_row, column=2)
        cell.value = indent + (line['product_id'][1].split(']')[-1].strip() if '][' in line['product_id'][1] else line['product_id'][1])
        cell.border = self.border
        cell.style = style
        
        # Cantidad
        cell = self.ws.cell(row=self.current_row, column=3)
        cell.value = qty
        cell.border = self.border
        cell.alignment = Alignment(horizontal="center")
        cell.style = style
        
        # Unidad
        cell = self.ws.cell(row=self.current_row, column=4)
        cell.value = line['product_uom_id'][1]
        cell.border = self.border
        cell.alignment = Alignment(horizontal="center")
        cell.style = style
        
        # Costo del Producto
        cell = self.ws.cell(row=self.current_row, column=5)
        cell.value = line['product_cost']
        cell.border = self.border
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = '"$"#,##0.00'
        cell.style = style
        
        # Costo de Materiales
        cell = self.ws.cell(row=self.current_row, column=6)
        cell.value = line['material_cost']
        cell.border = self.border
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = '"$"#,##0.00'
        cell.style = style
        
        self.current_row += 1
        
        # Si tiene sub-BOM, procesar recursivamente
        if line.get('sub_bom'):
            for sub_line in line['sub_bom']['lines']:
                self.write_bom_line(sub_line, qty)
    
    def write_operations(self, operations: List[Dict[str, Any]]):
        """
        Escribe las operaciones de fabricación
        Args:
            operations: Lista de operaciones
        """
        if not operations:
            return
            
        # Título de la sección
        self.current_row += 2
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = "Operaciones de fabricación"
        cell.font = Font(bold=True, size=12)
        self.current_row += 2
        
        # Encabezados
        headers = [
            "Secuencia",
            "Operación",
            "Centro de trabajo",
            "Tiempo (min)",
            "Eficiencia",
            "Costo/hora",
            "Costo Total"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.current_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Ajustar ancho de columnas para operaciones
        self.ws.column_dimensions['A'].width = 10  # Secuencia
        self.ws.column_dimensions['B'].width = 40  # Operación
        self.ws.column_dimensions['C'].width = 30  # Centro de trabajo
        self.ws.column_dimensions['D'].width = 15  # Tiempo
        self.ws.column_dimensions['E'].width = 12  # Eficiencia
        self.ws.column_dimensions['F'].width = 15  # Costo/hora
        self.ws.column_dimensions['G'].width = 15  # Costo Total
        
        self.current_row += 1
        total_operation_cost = 0.0
        
        # Datos de operaciones
        for op in operations:
            # Secuencia
            cell = self.ws.cell(row=self.current_row, column=1)
            cell.value = op['sequence']
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
            cell.style = self.operation_style
            
            # Operación
            cell = self.ws.cell(row=self.current_row, column=2)
            cell.value = op['name']
            cell.border = self.border
            cell.style = self.operation_style
            
            # Centro de trabajo
            cell = self.ws.cell(row=self.current_row, column=3)
            cell.value = op['workcenter_id'][1] if op['workcenter_id'] else ''
            cell.border = self.border
            cell.style = self.operation_style
            
            # Tiempo
            time_value = op['time_cycle_manual'] or op['time_cycle'] or 0.0
            cell = self.ws.cell(row=self.current_row, column=4)
            cell.value = time_value
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
            cell.style = self.operation_style
            cell.number_format = '#,##0.00'
            
            # Eficiencia
            cell = self.ws.cell(row=self.current_row, column=5)
            cell.value = op.get('efficiency', 1.0)
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
            cell.style = self.operation_style
            cell.number_format = '0%'
            
            # Costo por hora (del centro de trabajo)
            if op['workcenter_id']:
                cell = self.ws.cell(row=self.current_row, column=6)
                cell.value = float(op.get('costs_hour', 0.0))
                cell.border = self.border
                cell.alignment = Alignment(horizontal="right")
                cell.style = self.operation_style
                cell.number_format = '"$"#,##0.00'
            
            # Costo total de la operación
            cell = self.ws.cell(row=self.current_row, column=7)
            operation_cost = float(op.get('operation_cost', 0.0))
            cell.value = operation_cost
            cell.border = self.border
            cell.alignment = Alignment(horizontal="right")
            cell.style = self.operation_style
            cell.number_format = '"$"#,##0.00'
            
            total_operation_cost += operation_cost
            self.current_row += 1
        
        # Agregar total de costos de operaciones
        self.current_row += 1
        cell = self.ws.cell(row=self.current_row, column=6)
        cell.value = "Costo Total Operaciones:"
        cell.font = Font(bold=True)
        
        cell = self.ws.cell(row=self.current_row, column=7)
        cell.value = total_operation_cost
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        cell.number_format = '"$"#,##0.00'
    
    def export_bom(self, bom_data: Dict[str, Any], product_name: str) -> str:
        """
        Exporta los datos de la BOM a un archivo Excel
        Args:
            bom_data: Diccionario con los datos de la BOM
            product_name: Nombre del producto
        Returns:
            str: Ruta del archivo generado
        """
        try:
            # Información del producto
            cell = self.ws.cell(row=self.current_row, column=1)
            cell.value = "Información del producto"
            cell.font = Font(bold=True, size=12)
            self.current_row += 2
            
            # Sección de costos
            cell = self.ws.cell(row=self.current_row, column=1)
            cell.value = "Costos"
            cell.font = Font(bold=True)
            self.current_row += 1
            
            cost_data = [
                ("Costo del producto (standard_price):", f"${bom_data['product_cost']:.2f}"),
                ("Costo total de materiales:", f"${bom_data['total_material_cost']:.2f}"),
                ("Diferencia:", f"${(bom_data['product_cost'] - bom_data['total_material_cost']):.2f}")
            ]
            
            for label, value in cost_data:
                cell = self.ws.cell(row=self.current_row, column=1)
                cell.value = label
                cell.font = Font(bold=True)
                
                cell = self.ws.cell(row=self.current_row, column=2)
                cell.value = value
                if "Diferencia" in label:
                    diff = bom_data['product_cost'] - bom_data['total_material_cost']
                    if diff < 0:
                        cell.font = Font(color="FF0000")  # Rojo si es negativo
                    elif diff > 0:
                        cell.font = Font(color="008000")  # Verde si es positivo
                
                self.current_row += 1
            
            self.current_row += 1
            
            # Sección de rutas
            if bom_data['routes']:
                cell = self.ws.cell(row=self.current_row, column=1)
                cell.value = "Rutas de fabricación:"
                cell.font = Font(bold=True)
                
                cell = self.ws.cell(row=self.current_row, column=2)
                cell.value = ", ".join(bom_data['routes'])
                
                self.current_row += 2
            
            # Configurar encabezados de la BOM
            headers = ["Código", "Componente", "Cantidad", "Unidad", "Costo Producto", "Costo Materiales"]
            for col, header in enumerate(headers, start=1):
                cell = self.ws.cell(row=self.current_row, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
            
            self.current_row += 1
            
            # Ajustar ancho de columnas
            self.ws.column_dimensions['A'].width = 15  # Código
            self.ws.column_dimensions['B'].width = 60  # Componente (más ancho para indentación)
            self.ws.column_dimensions['C'].width = 12  # Cantidad
            self.ws.column_dimensions['D'].width = 12  # Unidad
            self.ws.column_dimensions['E'].width = 15  # Costo Producto
            self.ws.column_dimensions['F'].width = 15  # Costo Materiales
            
            # Procesar cada línea de la BOM
            for line in bom_data['lines']:
                self.write_bom_line(line)
            
            # Añadir operaciones
            if bom_data.get('operations'):
                self.write_operations(bom_data['operations'])
            
            # Crear directorio de exportación si no existe
            export_dir = "exports"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)
            
            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = "".join(c if c.isalnum() else "_" for c in product_name)
            filename = f"{export_dir}/BOM_{safe_name}_{timestamp}.xlsx"
            
            # Guardar archivo
            self.wb.save(filename)
            return filename
            
        except Exception as e:
            raise Exception(f"Error al exportar a Excel: {str(e)}") 