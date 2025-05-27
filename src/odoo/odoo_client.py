import odoorpc
from openpyxl import Workbook
from datetime import datetime
import re

class OdooClient:
    def __init__(self, url, db, username, password, port):
        # Asegurarse de que la URL tenga el formato correcto
        self.original_url = url
        self.host = self._clean_url(url)
        self.port = int(port)
        self.db = db
        self.username = username
        self.password = password
        # Para Odoo SaaS, siempre usamos HTTPS
        self.protocol = 'jsonrpc+ssl' if ('https://' in url or port == 443) else 'jsonrpc'
        self.odoo = self.connect()

    def _clean_url(self, url):
        """Limpia la URL eliminando el protocolo y la barra final si existe"""
        # Eliminar protocolo si existe
        url = re.sub(r'^https?://', '', url)
        # Eliminar barra final si existe
        url = url.rstrip('/')
        # Eliminar cualquier ruta después del dominio
        url = url.split('/')[0]
        return url

    def connect(self):
        """Establece la conexión con Odoo"""
        try:
            print(f"Intentando conectar a Odoo con:")
            print(f"URL original: {self.original_url}")
            print(f"Host limpio: {self.host}")
            print(f"Puerto: {self.port}")
            print(f"Protocolo: {self.protocol}")
            print(f"Base de datos: {self.db}")
            
            odoo = odoorpc.ODOO(
                self.host,
                port=self.port,
                protocol=self.protocol
            )
            odoo.login(self.db, self.username, self.password)
            return odoo
        except Exception as e:
            error_msg = str(e)
            if "HTTP Error 400" in error_msg:
                raise Exception(
                    "Error de conexión: Verifique que la URL y el puerto sean correctos.\n"
                    "Para Odoo SaaS (.odoo.com):\n"
                    "- URL: debe incluir 'https://' (ejemplo: 'https://mi-empresa.odoo.com')\n"
                    "- Puerto: debe ser 443\n"
                    "- Base de datos: generalmente es el nombre de su subdominio\n\n"
                    "Para Odoo local:\n"
                    "- URL: dominio o IP de su servidor\n"
                    "- Puerto: generalmente 8069\n"
                    f"\nError original: {error_msg}"
                )
            elif "SSL" in error_msg:
                raise Exception(
                    "Error SSL: La conexión requiere HTTPS.\n"
                    "- Asegúrese de que la URL comience con 'https://'\n"
                    "- Para Odoo SaaS, use el puerto 443\n"
                    f"\nError original: {error_msg}"
                )
            else:
                raise Exception(
                    f"Error de conexión: {error_msg}\n"
                    "Por favor verifique:\n"
                    "1. Que la URL sea correcta\n"
                    "2. Que el puerto sea correcto (443 para Odoo SaaS)\n"
                    "3. Que las credenciales sean correctas\n"
                    "4. Que la base de datos exista"
                )

    def get_products(self):
        """Obtiene la lista de productos que tienen lista de materiales"""
        try:
            # Obtener todos los BOMs
            Bom = self.odoo.env['mrp.bom']
            bom_ids = Bom.search([])
            
            if not bom_ids:
                return []
            
            # Obtener los BOMs y sus productos relacionados
            boms = Bom.browse(bom_ids)
            product_template_ids = []
            
            # Recolectar todos los IDs de plantillas de productos
            for bom in boms:
                product_template_ids.append(bom.product_tmpl_id.id)
            
            # Eliminar duplicados
            product_template_ids = list(set(product_template_ids))
            
            # Obtener los productos relacionados con estas plantillas
            Product = self.odoo.env['product.product']
            products = []
            
            for template_id in product_template_ids:
                # Buscar variantes de producto para esta plantilla
                variant_ids = Product.search([('product_tmpl_id', '=', template_id)])
                for product in Product.browse(variant_ids):
                    products.append({
                        'id': product.id,
                        'name': f"[{product.default_code}] {product.name}" if product.default_code else product.name
                    })
            
            return sorted(products, key=lambda x: x['name'])
        except Exception as e:
            print(f"Error detallado al obtener productos: {str(e)}")
            raise Exception(f"Error al cargar productos: {str(e)}")

    def get_bom_data(self, product_id):
        """Obtiene los datos de la lista de materiales para un producto"""
        Product = self.odoo.env['product.product']
        Bom = self.odoo.env['mrp.bom']
        
        product = Product.browse(product_id)
        bom = Bom.search([('product_tmpl_id', '=', product.product_tmpl_id.id)], limit=1)
        
        if not bom:
            raise Exception("No se encontró lista de materiales para este producto")
            
        bom_data = []
        bom = Bom.browse(bom[0])
        
        for line in bom.bom_line_ids:
            component = line.product_id
            route_name = ''
            if component.route_ids:
                route = self.odoo.env['stock.route'].browse(component.route_ids[0])
                route_name = route.name
                
            bom_data.append({
                'Producto': f"[{component.default_code}] {component.name}" if component.default_code else component.name,
                'Cantidad': line.product_qty,
                'Plazo de entrega': component.produce_delay or 0,
                'Ruta': route_name,
                'Costo BOM': self.get_bom_cost(component.id),
                'Costo Producto': component.standard_price
            })
            
        return bom_data

    def get_bom_cost(self, product_id):
        """Calcula el costo total de la lista de materiales de un producto"""
        Bom = self.odoo.env['mrp.bom']
        bom = Bom.search([('product_id', '=', product_id)], limit=1)
        
        if not bom:
            return 0.0
            
        total_cost = 0.0
        bom = Bom.browse(bom[0])
        
        for line in bom.bom_line_ids:
            component_cost = line.product_id.standard_price
            component_bom_cost = self.get_bom_cost(line.product_id.id)
            total_cost += (component_cost + component_bom_cost) * line.product_qty
            
        return total_cost

    def export_bom(self, product_id):
        """Exporta la lista de materiales a un archivo Excel"""
        bom_data = self.get_bom_data(product_id)
        product = self.odoo.env['product.product'].browse(product_id)
        
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Materiales"
        
        # Escribir encabezados
        headers = ['Producto', 'Cantidad', 'Plazo de entrega', 'Ruta', 'Costo BOM', 'Costo Producto']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Escribir datos
        for row, item in enumerate(bom_data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item[header])
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Crear nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BOM_{product.default_code or 'producto'}_{timestamp}.xlsx"
        
        # Guardar archivo
        wb.save(filename)
        return filename 